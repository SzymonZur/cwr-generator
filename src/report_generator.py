"""XLSX report generator for Creative Work Reports."""

import logging
from pathlib import Path
from typing import Dict, Any, List, Optional
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Generate XLSX reports from template."""

    def __init__(self, template_path: Path):
        """
        Initialize report generator.

        Args:
            template_path: Path to Excel template file
        """
        self.template_path = template_path
        self.wb = None
        self.ws = None

    def load_template(self):
        """Load the Excel template."""
        logger.info(f"Loading template from {self.template_path}")
        self.wb = openpyxl.load_workbook(self.template_path)
        self.ws = self.wb["CreativeTime"]
        logger.info("Template loaded successfully")

    def populate_header(self, employee_name: str, company_name: str, year: int):
        """
        Populate the header section of the report.

        Args:
            employee_name: Name of the employee
            company_name: Company name
            year: Year for the report
        """
        logger.info("Populating header section")

        # Row 2: Employee Name
        self.ws.cell(row=2, column=1, value=employee_name)

        # Row 3: Company Name
        self.ws.cell(row=3, column=1, value=company_name)

        # Row 4: Report Period
        start_date = datetime(year, 1, 1)
        end_date = datetime(year, 12, 31)

        # Format: DD MM YYYY
        start_str = start_date.strftime("%d %m %Y")
        end_str = end_date.strftime("%d %m %Y")

        self.ws.cell(row=4, column=4, value=start_str)
        self.ws.cell(row=4, column=6, value=end_str)

    def find_insertion_row(self) -> int:
        """
        Find the row where new project entries should be inserted.
        The template has pre-styled rows starting at row 7 (each project takes 3 rows).

        Returns:
            Row number to start filling projects
        """
        # Template has pre-styled project rows starting at row 7
        # Each project takes 3 rows (row, row+1, row+2)
        # Look for the first row with a number in column A that we should fill
        # We'll fill existing template rows, starting from row 7

        current_row = 7
        max_search_rows = 200  # Reasonable limit to avoid searching entire sheet

        # Find the first row that has a number (template example) or is empty
        # We'll start filling from row 7 (first template project row)
        while current_row <= min(self.ws.max_row, max_search_rows):
            cell_value = self.ws.cell(row=current_row, column=1).value
            
            # Check if this is a project row (has a number) or empty styled row
            # We'll fill it regardless - template rows are meant to be filled
            if cell_value is None or isinstance(cell_value, (int, float)):
                # This is a valid project row to fill
                logger.info(f"Will fill projects starting at row {current_row}")
                return current_row
            
            current_row += 3  # Each project takes 3 rows

        # If we didn't find a suitable row, start at row 7
        logger.info(f"Will fill projects starting at row 7")
        return 7

    def _unmerge_cells_in_range(
        self, start_row: int, end_row: int, start_col: int, end_col: int
    ):
        """Unmerge cells in the specified range."""
        ranges_to_remove = []
        for merged_range in list(self.ws.merged_cells.ranges):
            # Check if merged range overlaps with our target range
            if (
                merged_range.min_row <= end_row
                and merged_range.max_row >= start_row
                and merged_range.min_col <= end_col
                and merged_range.max_col >= start_col
            ):
                ranges_to_remove.append(merged_range)

        for merged_range in ranges_to_remove:
            self.ws.unmerge_cells(str(merged_range))

    def fill_project(
        self,
        row: int,
        project_number: float,
        project_name: str,
        creative_work_details: str,
        contracted_time: Optional[float] = None,
        non_creative_time: Optional[float] = None,
        ctd_allocation: Optional[float] = None,
    ):
        """
        Fill a project entry into the pre-styled template rows.

        Args:
            row: Starting row number for the project entry (first of 3 rows)
            project_number: Project number (1.0, 2.0, etc.)
            project_name: Name of the project
            creative_work_details: Description of creative work
            contracted_time: Contracted time in hours (optional)
            non_creative_time: Non-creative time in hours (optional)
            ctd_allocation: CTD allocation percentage (optional)
        """
        logger.debug(f"Filling project {project_number} at row {row}")

        # Unmerge any existing merged cells in the project rows (rows row to row+2, cols 1-7)
        # This is necessary because merged cells (except the top-left) are read-only
        self._unmerge_cells_in_range(row, row + 2, 1, 7)

        # Template structure (based on row 6 headers):
        # Column A: Number
        # Column B: Project Name  
        # Column C: Creative Works Details
        # Column D: LIABILITY CALCULATION (labels)
        # Column E: Values (hours, percentages)
        # Column F: Final Liability (formula)
        # Column G: Approver / Date of Approval

        # Row 1 of project entry (main row) - row
        self.ws.cell(row=row, column=1, value=project_number)  # A: Number
        self.ws.cell(row=row, column=2, value=project_name)  # B: Project Name
        self.ws.cell(row=row, column=3, value=creative_work_details)  # C: Creative Works Details
        self.ws.cell(row=row, column=4, value="Contracted Time for Project")  # D: Label
        if contracted_time is not None:
            self.ws.cell(row=row, column=5, value=contracted_time)  # E: Value
        # F: Formula for final liability
        formula_row = row
        self.ws.cell(
            row=row,
            column=6,
            value=f"=(E{formula_row}-E{formula_row+1})*E{formula_row+2}",
        )  # F: Final Liability formula
        self.ws.cell(row=row, column=7, value="Name of Approver")  # G: Approver label

        # Row 2 of project entry - row + 1
        # Columns A, B, C are merged (empty)
        self.ws.cell(row=row + 1, column=4, value="Non-Creative Time Spent on Project")  # D: Label
        if non_creative_time is not None:
            self.ws.cell(row=row + 1, column=5, value=non_creative_time)  # E: Value
        # Column F is merged (empty)
        # Column G is empty

        # Row 3 of project entry - row + 2
        # Columns A, B, C are merged (empty)
        self.ws.cell(row=row + 2, column=4, value="CTD allocation per EA Addendum")  # D: Label
        if ctd_allocation is not None:
            self.ws.cell(row=row + 2, column=5, value=ctd_allocation)  # E: Value
        # Column F is merged (empty)
        self.ws.cell(row=row + 2, column=7, value="Double click for Date")  # G: Date label

        # Ensure cells are merged to match template structure
        # A, B, C, and F should be merged across all 3 rows
        try:
            self.ws.merge_cells(f"A{row}:A{row+2}")  # Merge column A
        except ValueError:
            pass  # Already merged
        
        try:
            self.ws.merge_cells(f"B{row}:B{row+2}")  # Merge column B
        except ValueError:
            pass  # Already merged
        
        try:
            self.ws.merge_cells(f"C{row}:C{row+2}")  # Merge column C
        except ValueError:
            pass  # Already merged
        
        try:
            self.ws.merge_cells(f"F{row}:F{row+2}")  # Merge column F (formula)
        except ValueError:
            pass  # Already merged

        # Set text wrapping and alignment for the creative work details cell (column C)
        creative_cell = self.ws.cell(row=row, column=3)
        creative_cell.alignment = Alignment(
            wrap_text=True, vertical="top", horizontal="left"
        )

        # Set alignment for project name (column B)
        name_cell = self.ws.cell(row=row, column=2)
        name_cell.alignment = Alignment(
            wrap_text=True, vertical="top", horizontal="left"
        )

    def generate_report(
        self,
        employee_name: str,
        company_name: str,
        year: int,
        projects: Dict[str, Dict[str, Any]],
        project_summaries: Dict[str, Dict[str, str]],
    ) -> openpyxl.Workbook:
        """
        Generate the complete report.

        Args:
            employee_name: Name of the employee
            company_name: Company name
            year: Year for the report
            projects: Dictionary of project data (from DataProcessor)
            project_summaries: Dictionary mapping project keys to AI summaries

        Returns:
            The workbook object
        """
        logger.info("Generating report")

        self.load_template()
        self.populate_header(employee_name, company_name, year)

        # Find where to insert projects
        insertion_row = self.find_insertion_row()

        # Sort projects by key (excluding UNLINKED projects)
        sorted_projects = sorted(
            [(k, v) for k, v in projects.items() if not k.startswith("UNLINKED")],
            key=lambda x: x[0],
        )

        # Add unlinked projects at the end
        unlinked_projects = [
            (k, v) for k, v in projects.items() if k.startswith("UNLINKED")
        ]
        sorted_projects.extend(unlinked_projects)

        # Insert each project
        current_row = insertion_row
        project_number = 1.0

        for project_key, project_data in sorted_projects:
            # Get summary
            summary = project_summaries.get(project_key, {})

            project_name = project_data.get("project_name", project_key)
            creative_work_details = summary.get(
                "creative_work_details", summary.get("description", "Development work")
            )

            # Get metrics for time estimation (optional)
            metrics = project_data.get("metrics", {})
            total_commits = metrics.get("total_commits", 0)

            # Estimate time based on commits (rough estimate: 2-4 hours per commit)
            # This is a placeholder - actual time should come from Jira or manual input
            estimated_hours = total_commits * 3 if total_commits > 0 else None

            self.fill_project(
                row=current_row,
                project_number=project_number,
                project_name=project_name,
                creative_work_details=creative_work_details,
                contracted_time=estimated_hours,
                non_creative_time=None,  # Leave for manual entry
                ctd_allocation=0.75,  # Default value, can be customized
            )

            current_row += 3  # Each project takes 3 rows
            project_number += 1.0

        logger.info(f"Inserted {len(sorted_projects)} projects into report")
        return self.wb

    def save_report(self, output_path: Path):
        """
        Save the report to a file.

        Args:
            output_path: Path where to save the report
        """
        if self.wb is None:
            raise ValueError("Report not generated. Call generate_report() first.")

        logger.info(f"Saving report to {output_path}")
        self.wb.save(output_path)
        logger.info("Report saved successfully")
