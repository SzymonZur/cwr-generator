"""Tests for XLSX report generator."""

import unittest
import tempfile
from pathlib import Path
from datetime import datetime
from src.report_generator import ReportGenerator


class TestReportGenerator(unittest.TestCase):
    """Test report generator."""

    def setUp(self):
        """Set up test fixtures."""
        # Use the actual template
        template_path = (
            Path(__file__).parent.parent
            / "templates"
            / "PL_time_report.xlsx"
        )
        if template_path.exists():
            self.template_path = template_path
        else:
            self.skipTest("Template file not found")

    def test_load_template(self):
        """Test loading the template."""
        generator = ReportGenerator(self.template_path)
        generator.load_template()

        self.assertIsNotNone(generator.wb)
        self.assertIsNotNone(generator.ws)
        self.assertEqual(generator.ws.title, "CreativeTime")

    def test_populate_header(self):
        """Test populating header section."""
        generator = ReportGenerator(self.template_path)
        generator.load_template()

        generator.populate_header(
            employee_name="Test User", company_name="Test Company", year=2024
        )

        # Check employee name
        self.assertEqual(generator.ws.cell(row=2, column=1).value, "Test User")

        # Check company name
        self.assertEqual(generator.ws.cell(row=3, column=1).value, "Test Company")

        # Check dates
        start_date = generator.ws.cell(row=4, column=4).value
        end_date = generator.ws.cell(row=4, column=6).value

        self.assertIsNotNone(start_date)
        self.assertIsNotNone(end_date)
        self.assertIn("2024", str(start_date))
        self.assertIn("2024", str(end_date))

    def test_insert_project(self):
        """Test filling a project entry."""
        generator = ReportGenerator(self.template_path)
        generator.load_template()

        generator.fill_project(
            row=10,
            project_number=1.0,
            project_name="Test Project",
            creative_work_details="Test description",
            contracted_time=100.0,
            non_creative_time=20.0,
            ctd_allocation=0.75,
        )

        # Check project number
        self.assertEqual(generator.ws.cell(row=10, column=1).value, 1.0)

        # Check project name
        self.assertEqual(generator.ws.cell(row=10, column=2).value, "Test Project")

        # Check description
        self.assertEqual(generator.ws.cell(row=10, column=3).value, "Test description")

        # Check contracted time
        self.assertEqual(generator.ws.cell(row=10, column=5).value, 100.0)

        # Check non-creative time
        self.assertEqual(generator.ws.cell(row=11, column=5).value, 20.0)

        # Check CTD allocation
        self.assertEqual(generator.ws.cell(row=12, column=5).value, 0.75)

        # Check formula
        formula = generator.ws.cell(row=10, column=6).value
        self.assertIn("=(E10-E11)*E12", str(formula))

    def test_generate_and_save_report(self):
        """Test generating and saving a complete report."""
        generator = ReportGenerator(self.template_path)

        projects = {
            "FUI": {
                "project_key": "FUI",
                "project_name": "Feature UI",
                "commits": [],
                "tickets": {},
                "metrics": {"total_commits": 5},
            }
        }

        project_summaries = {
            "FUI": {
                "description": "UI Feature Project",
                "creative_work_details": "Developed new UI features",
                "technical_summary": "React components",
            }
        }

        generator.generate_report(
            employee_name="Test User",
            company_name="Test Company",
            year=2024,
            projects=projects,
            project_summaries=project_summaries,
        )

        # Save to temporary file
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = Path(tmp.name)

        try:
            generator.save_report(output_path)

            # Verify file was created
            self.assertTrue(output_path.exists())

            # Try to load it back
            import openpyxl

            wb = openpyxl.load_workbook(output_path)
            ws = wb["CreativeTime"]

            # Check header
            self.assertEqual(ws.cell(row=2, column=1).value, "Test User")
            self.assertEqual(ws.cell(row=3, column=1).value, "Test Company")

        finally:
            # Clean up
            if output_path.exists():
                output_path.unlink()


if __name__ == "__main__":
    unittest.main()
